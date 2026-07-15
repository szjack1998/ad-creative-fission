#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
强哥广告skill · 脚本裂变工具 fission_tool.py
============================================
自动流程：读 Excel 脚本库 -> 读视频表现 CSV -> 命名匹配 -> 指标归因
        -> 基于 winning 母本 + 模板池生成裂变变体 -> 输出 HTML 报告

用法:
  python fission_tool.py --scripts 脚本库.xlsx --videos 视频数据.csv --output 报告.html

可选:
  --winner-yg / --winner-yyy / --winner 数字人 / --winner 图文   手动指定母本系列 (默认自动选最高CVR)
  --top-n 18                                                   裂变脚本条数 (默认18)

CSV 格式约定 (列名不区分大小写, 自动识别):
  name, date, spend, ctr, cvr        # cvr 也可用 download_rate / conversion_rate 列名
  例:
    0415-新程教育-yg11,06-18,226,1.22,7.41
    6_22-新程教育-yyy-3,06-22,4109,0.97,33.31

依赖: openpyxl (读xlsx) + 标准库 csv/html。
  pip install openpyxl
"""

import argparse
import csv
import html
import os
import re
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("缺少依赖 openpyxl，请先: pip install openpyxl")


# ============================================================
# 1. 读取脚本库 (Excel)
# ============================================================
def read_scripts(xlsx_path):
    """读取脚本库 xlsx，返回 [(row_idx, text)] 列表。
    结构约定: 第二列(索引1)为'文案'，第一列为空或序号。"""
    SKIP_MARKERS = ("更新", "日更", "续更", "待补", "注：", "备注", "说明")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    scripts = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row,
                                          max_col=ws.max_column, values_only=True), start=1):
        # 取第二列文案，fallback 取第一个非空单元格
        text = row[1] if len(row) > 1 else None
        if not text:
            text = next((c for c in row if c), None)
        stripped = str(text).strip() if text else ""
        if len(stripped) < 2:
            continue
        # 跳过短标注行（如「7.3日更新」）
        if len(stripped) < 12 and any(m in stripped for m in SKIP_MARKERS):
            continue
        scripts.append((i, stripped))
    return scripts


def classify_script(text):
    """按内容关键词把脚本分类为 yg/yyy/数字人/通用/大字报。
    返回 (series, persona)"""
    t = text
    # 大字报: 多行 / 含换行且短促 / 含"最后窗口""最容易考"等
    if "\n" in t and len(t) < 120:
        series = "大字报"
    elif any(k in t for k in ["宝妈", "女人", "40岁", "女性"]):
        series = "yyy"  # 特定人群锚定型
    elif any(k in t for k in ["逼着自己", "别再刷", "哭着回来", "错过", "后悔", "最后", "抢"]):
        series = "yg"   # 好奇/紧迫/反直觉
    elif any(k in t for k in ["不限专业", "转型", "转行", "副业", "接单"]):
        series = "yg"
    else:
        series = "通用"
    # 人群
    if "宝妈" in t:
        persona = "宝妈"
    elif "40岁" in t or "女人" in t:
        persona = "女性40+"
    elif "转行" in t or "转型" in t:
        persona = "转行族"
    elif "副业" in t or "接单" in t:
        persona = "副业族"
    else:
        persona = "通用"
    return series, persona


# ============================================================
# 2. 读取视频数据 (CSV)
# ============================================================
def read_videos(csv_path):
    """读取视频表现 CSV，自动识别列名。返回 list[dict]。"""
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if not rows:
        return []
    # 列名归一化
    norm = {}
    for k in rows[0].keys():
        nk = k.strip().lower()
        if nk in ("name", "视频", "素材", "创意"):
            norm[k] = "name"
        elif nk in ("date", "创意时间", "时间"):
            norm[k] = "date"
        elif nk in ("spend", "花费", "消耗"):
            norm[k] = "spend"
        elif nk in ("ctr", "点击率"):
            norm[k] = "ctr"
        elif nk in ("cvr", "download_rate", "download rate", "下载率", "转化", "conversion", "conversion_rate"):
            norm[k] = "cvr"
    out = []
    for r in rows:
        rec = {}
        for k, v in r.items():
            if k in norm:
                rec[norm[k]] = v
        rec["name"] = rec.get("name", "")
        rec["spend"] = float(rec.get("spend", 0) or 0)
        rec["ctr"] = float(rec.get("ctr", 0) or 0)
        rec["cvr"] = float(rec.get("cvr", 0) or 0)
        rec["date"] = rec.get("date", "")
        out.append(rec)
    return out


def detect_series(name):
    """从视频名检测系列: yg / yyy / 数字人 / 图文 / 未知"""
    n = name.lower()
    if "数字人" in name or "数智人" in name:
        return "数字人"
    if "图文" in name:
        return "图文"
    # yy / yg 系列: 取最靠近的标记
    m = re.search(r"(yg|yyy|y|y0{0,2})[-_]?\d*", n)
    if m:
        token = m.group(1)
        if token.startswith("yg") and "yyy" not in token:
            return "yg"
        if "yyy" in token or token == "yy":
            return "yyy"
    if "yg" in n:
        return "yg"
    if "yyy" in n:
        return "yyy"
    return "未知"


# ============================================================
# 3. 匹配 + 归因
# ============================================================
def match_and_analyze(videos, scripts):
    """把视频按命名匹配到脚本类别，并计算归因指标。"""
    # 分类脚本
    by_series = {"yg": [], "yyy": [], "数字人": [], "图文": [], "通用": [], "大字报": []}
    for idx, text in scripts:
        series, persona = classify_script(text)
        by_series.setdefault(series, []).append((idx, text, persona))

    matches = []
    for v in videos:
        s = detect_series(v["name"])
        # 找同类脚本
        pool = by_series.get(s, [])
        if not pool:
            pool = by_series.get("通用", [])
        matched = "; ".join(f"行{r}" for r, _, _ in pool[:3]) or "（无对应脚本）"
        conf = "高匹配" if pool else "低匹配"
        v["series"] = s
        v["matched"] = matched
        v["conf"] = conf
        matches.append(v)

    # 归因: 排名 / 效率王 / 零转化 / CPM估算
    total_spend = sum(v["spend"] for v in videos) or 1
    for v in videos:
        v["spend_pct"] = round(v["spend"] / total_spend * 100, 1)
    # 效率王: 排除预算黑洞(花费占比>35%)，在剩余里取CVR最高；无CVR则取CTR最高
    candidates = [v for v in videos if v["spend_pct"] <= 35] or videos
    with_cvr = [v for v in candidates if v["cvr"] > 0]
    winner = (max(with_cvr, key=lambda x: x["cvr"]) if with_cvr
              else max(candidates, key=lambda x: x["ctr"])) if candidates else None
    zero_cvr = [v for v in videos if v["cvr"] == 0]
    high_ctr_zero = [v for v in videos if v["ctr"] >= 1.5 and v["cvr"] == 0]
    return matches, by_series, winner, zero_cvr, high_ctr_zero, total_spend


# ============================================================
# 4. 裂变生成 (模板池 + winning 母本)
# ============================================================
HOOK_POOL = {
    "行为中断": "别再刷手机了！你刷的这10分钟，够你了解一个能让收入翻倍的证书。",
    "社交证明": "我同事去年偷偷考了个证，现在工资涨了3000块，每天准时下班。",
    "FOMO窗口": "听句劝：健康管理师这个证，7月这场是你最容易过的一次，后面难度只会越来越高。",
    "反常识": "逼着自己去考个健康管理师证，两个月后你会哭着回来感谢自己。",
    "错失恐惧": "2026年错过就会后悔的证书，现在知道的人不多，但门槛马上就要涨了。",
    "机会成本": "你每天刷手机的时间加起来至少2小时，如果把其中1小时用来备考健康管理师……",
    "年龄焦虑": "35岁以后最怕什么？不是失业，是想转行时发现自己没有任何能拿得出手的证书。",
    "行业缺口": "国家卫健委统计：健康管理师行业缺口超100万，现在考了就不愁找不到好工作。",
}
PERSONA_POOL = {
    "通用": "",
    "宝妈": "很多宝妈利用孩子睡觉的1-2小时就考下了健康管理师，不用坐班、手机接单、时间自由。",
    "打工人": "加班到10点回家累得不想说话？有些人已经悄悄考下证，现在朝九晚五周末双休。",
    "转行族": "想转型健康行业？健康管理师是不限专业不限经验就能跨行的跳板。",
    "副业族": "不用辞职、不用坐班，考个证手机上就能接健康咨询单子，一条回复几十到几百。",
}
BENEFIT_POOL = [
    "只考选择题、60分及格、不限专业就能报",
    "拿证后学校/社区/事业单位抢着要，周末双休五险一金",
    "居家也能接单赚钱，时间完全自由",
    "2026年是最后一年低难度窗口期，往后只会更难",
]
CTA_POOL = [
    "想知道你现在能不能直接报名？来直播间测一测。",
    "来直播间自查一下你符不符合条件。",
    "别等难度涨了才后悔，来直播间看你还赶不赶上。",
    "想了解怎么接单？来直播间。",
]

PRIORITY_PLAN = [
    ("P0", 4, "紧急生产"), ("P1", 6, "A/B测试"),
    ("P2", 5, "边界探索"), ("大字报", 3, "大字报短句"),
]


def generate_fission(winner_series, n=18):
    """基于 winning 母本系列 + 模板池，生成 n 条裂变变体。
    返回 [dict(tag, priority, hook, body, vars, hypothesis)]"""
    results = []
    hooks = list(HOOK_POOL.items())
    personas = list(PERSONA_POOL.items())
    prios = []
    for p, cnt, _ in PRIORITY_PLAN:
        prios += [p] * cnt
    prios = prios[:n]

    for i in range(n):
        p = prios[i] if i < len(prios) else "P2"
        hook_name, hook = hooks[i % len(hooks)]
        pers_name, pers = personas[(i // 2) % len(personas)]
        ben = BENEFIT_POOL[i % len(BENEFIT_POOL)]
        cta = CTA_POOL[i % len(CTA_POOL)]

        if p == "大字报":
            body = (f"<b>健康管理师证书</b><br>━━━━━━━━━━━<br>"
                    f"{hook_name}窗口<br>{ben.split('，')[0]}<br>"
                    f"{ben.split('，')[1] if '，' in ben else ''}<br>"
                    f"🔥 来直播间 抢位置")
            vars_ = ["大字报", "短句", hook_name, "视觉冲击"]
        else:
            ptext = f" {pers}" if pers else ""
            body = f"{hook}{ptext}健康管理师证书，{ben}。{cta}"
            vars_ = [hook_name, pers_name or "通用", "3硬利益", "条件CTA"]

        hypo = (f"以 {winner_series} 母本结构验证「{hook_name}」钩子在"
                f"{pers_name or '通用'}人群的 CVR 稳定性")
        results.append({
            "tag": f"[{p}-{i+1}]", "priority": p,
            "hook": hook, "body": body, "vars": vars_, "hypothesis": hypo,
        })
    return results


# ============================================================
# 5. 渲染 HTML
# ============================================================
def render_html(videos, by_series, winner, zero_cvr, high_ctr_zero,
                fission, scripts_count, out_path):
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    # 匹配表
    rows_html = ""
    for v in videos:
        color = "#3B6D11" if v["cvr"] > 0 else "#E24B4A"
        rows_html += f"""<tr>
<td><b>{html.escape(v['name'])}</b><br><small style="color:#888">{html.escape(v.get('date',''))}</small></td>
<td><span class="tag tag-{ 'yg' if v['series']=='yg' else ('yyy' if v['series']=='yyy' else ('dzr' if v['series']=='数字人' else 'p2')) }">{v['series']}</span></td>
<td>¥{v['spend']:,.0f}</td>
<td>{v['ctr']}%</td>
<td style="color:{color};font-weight:500">{v['cvr']}%</td>
<td>{v['spend_pct']}%</td>
<td>{html.escape(v['matched'])}</td>
<td><span class="mb">{v['conf']}</span></td>
</tr>"""

    # 裂变表
    fission_html = ""
    for f in fission:
        cls = {"P0": "p0", "P1": "p1", "P2": "p2", "大字报": "p1"}.get(f["priority"], "p2")
        tags = "".join(f'<span class="vt">{html.escape(x)}</span>' for x in f["vars"] if x)
        fission_html += f"""<div class="sc {cls}">
<span class="hk">{html.escape(f['tag'])} {html.escape(f['hook'][:24])}…</span>
<div class="bd">{f['body']}</div>
<div class="mt">{tags}</div>
</div>"""

    total_spend = sum(v["spend"] for v in videos)
    winner_line = (f"{html.escape(winner['name'])} (CVR {winner['cvr']}%)" if winner else "无")
    zero_line = ", ".join(html.escape(v["name"]) for v in zero_cvr) or "无"
    alert_line = ", ".join(html.escape(v["name"]) for v in high_ctr_zero) or "无"

    html_doc = f"""<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>强哥广告 · 裂变报告</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,"PingFang SC","Microsoft YaHei",sans-serif;background:#F8F8F7;color:#2C2C2A;line-height:1.6;font-size:13px;padding:20px;max-width:1100px;margin:0 auto}}
h1{{font-size:20px;font-weight:600;margin-bottom:4px}}
h2{{font-size:15px;font-weight:600;border-left:3px solid #534AB7;padding-left:10px;margin:24px 0 12px}}
.sub{{color:#888;font-size:12px;margin-bottom:16px}}
.card{{background:#fff;border-radius:12px;border:.5px solid rgba(44,44,42,.1);padding:18px;margin-bottom:14px}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
th{{background:#F8F8F7;font-weight:500;text-align:left;padding:7px 9px;border-bottom:1px solid rgba(44,44,42,.18)}}
td{{padding:8px 9px;border-bottom:.5px solid rgba(44,44,42,.1);vertical-align:top}}
.tag{{display:inline-block;padding:1px 8px;border-radius:10px;font-size:11px}}
.tag-yg{{background:#EEEDFE;color:#534AB7}} .tag-yyy{{background:#FAECE7;color:#D85A30}}
.tag-dzr{{background:#E1F5EE;color:#0F6E56}} .tag-p2{{background:#E6F1FB;color:#185FA5}}
.mb{{font-size:11px;padding:1px 7px;border-radius:6px;background:#f0f0ee;color:#888}}
.metrics{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:4px}}
.metric{{background:#F8F8F7;border-radius:8px;padding:12px;text-align:center}}
.metric .v{{font-size:22px;font-weight:600}} .metric .l{{font-size:11px;color:#888;margin-top:2px}}
.fission{{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:10px}}
.sc{{background:#F8F8F7;border-radius:8px;padding:12px;border-left:3px solid rgba(44,44,42,.18)}}
.sc.p0{{border-left-color:#E24B4A}} .sc.p1{{border-left-color:#BA7517}} .sc.p2{{border-left-color:#185FA5}}
.hk{{font-weight:500;color:#534AB7;font-size:12px;display:block;margin-bottom:5px}}
.bd{{color:#5F5E5A;font-size:12px;line-height:1.7}}
.mt{{margin-top:7px;display:flex;gap:5px;flex-wrap:wrap}} .vt{{font-size:10px;background:#fff;color:#5F5E5A;border:.5px solid rgba(44,44,42,.15);border-radius:4px;padding:1px 6px}}
.hl{{border-radius:8px;padding:12px 14px;margin:8px 0;font-size:12px}}
.hl.r{{background:#FCEBEB;color:#E24B4A}} .hl.g{{background:#EAF3DE;color:#3B6D11}} .hl.a{{background:#FAEEDA;color:#BA7517}}
@media(max-width:700px){{.metrics{{grid-template-columns:repeat(2,1fr)}}.fission{{grid-template-columns:1fr}}}}
</style></head><body>
<h1>强哥广告 · 脚本裂变报告</h1>
<p class="sub">自动生成 · {now} · 数据源: 视频CSV × Excel脚本库({scripts_count}条)</p>

<div class="card"><div class="metrics">
<div class="metric"><div class="v" style="color:#534AB7">{len(videos)}</div><div class="l">视频素材</div></div>
<div class="metric"><div class="v" style="color:#534AB7">{scripts_count}</div><div class="l">脚本库条数</div></div>
<div class="metric"><div class="v">¥{total_spend:,.0f}</div><div class="l">总花费</div></div>
<div class="metric"><div class="v" style="color:#E24B4A">{len(fission)}</div><div class="l">裂变脚本</div></div>
</div></div>

<h2>一、视频-脚本匹配 & 归因</h2>
<div class="card"><table>
<thead><tr><th>视频</th><th>系列</th><th>花费</th><th>CTR</th><th>CVR</th><th>占比</th><th>匹配脚本</th><th>置信</th></tr></thead>
<tbody>{rows_html}</tbody></table></div>

<div class="hl g"><b>效率王 (winning母本):</b> {winner_line}</div>
<div class="hl r"><b>零转化素材:</b> {zero_line}</div>
<div class="hl a"><b>高点击零转化(人群错配嫌疑):</b> {alert_line}</div>

<h2>二、裂变脚本矩阵 (自动生成 · 需人工润色/实测)</h2>
<div class="hl a">以下变体由「winning母本结构 + 钩子/人群/利益模板池」规则生成，作为创作起点；最终文案请结合品牌调性润色，并小预算实测验证。</div>
<div class="card"><div class="fission">{fission_html}</div></div>

<div style="text-align:center;padding:24px 0;color:#888;font-size:11px">强哥广告skill · fission_tool.py · 自动生成</div>
</body></html>"""
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html_doc)
    return out_path


# ============================================================
# 主流程
# ============================================================
def main():
    ap = argparse.ArgumentParser(description="强哥广告脚本裂变工具")
    ap.add_argument("--scripts", required=True, help="脚本库 xlsx 路径")
    ap.add_argument("--videos", required=True, help="视频表现 CSV 路径")
    ap.add_argument("--output", default="裂变报告.html", help="输出 HTML 路径")
    ap.add_argument("--winner", default=None, help="手动指定母本系列: yg/yyy/数字人/图文")
    ap.add_argument("--top-n", type=int, default=18, help="裂变脚本条数")
    args = ap.parse_args()

    print(f"[1/5] 读取脚本库: {args.scripts}")
    scripts = read_scripts(args.scripts)
    print(f"      共 {len(scripts)} 条脚本")

    print(f"[2/5] 读取视频数据: {args.videos}")
    videos = read_videos(args.videos)
    print(f"      共 {len(videos)} 条视频")

    print("[3/5] 命名匹配 + 指标归因")
    matches, by_series, winner, zero_cvr, high_ctr_zero, _ = match_and_analyze(videos, scripts)
    if args.winner:
        winner_series = args.winner
    elif winner:
        winner_series = winner["series"]
    else:
        winner_series = "yg"
    print(f"      效率王: {winner['name'] if winner else '无'} | 零转化: {len(zero_cvr)} 条")

    print(f"[4/5] 生成裂解变体 (母本={winner_series}, n={args.top_n})")
    fission = generate_fission(winner_series, args.top_n)
    print(f"      生成 {len(fission)} 条")

    print(f"[5/5] 渲染 HTML: {args.output}")
    render_html(videos, by_series, winner, zero_cvr, high_ctr_zero,
                fission, len(scripts), args.output)
    print("完成 ✓")
    print(f"  报告: {os.path.abspath(args.output)}")


if __name__ == "__main__":
    main()
